"""Document converters — PDF, DOCX, XLSX, images → markdown."""

from __future__ import annotations

import asyncio
import io
import json
import logging
import subprocess
import tempfile
from pathlib import Path
from typing import Any

from PIL import Image
import pytesseract

logger = logging.getLogger(__name__)


# ── PDF → markdown + page images ─────────────────────────────────────────────


async def pdf_to_package(
    content: bytes,
) -> tuple[list[bytes], str, list[dict[str, Any]]]:
    """Convert a PDF to page images, a combined markdown string, and a page map.

    Uses ``marker`` for high-quality conversion.  Falls back to a simpler
    approach if marker is unavailable.

    Returns:
        (page_images, combined_markdown, page_map)
        page_map is a list of {page: int, char_start: int, char_end: int}.
    """
    try:
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(None, _pdf_to_package_sync, content)
    except Exception:
        logger.error("PDF conversion failed entirely", exc_info=True)
        return [], "(PDF conversion failed)", []


def _pdf_to_package_sync(
    content: bytes,
) -> tuple[list[bytes], str, list[dict[str, Any]]]:
    with tempfile.TemporaryDirectory() as tmpdir:
        pdf_path = Path(tmpdir) / "input.pdf"
        pdf_path.write_bytes(content)
        out_dir = Path(tmpdir) / "output"
        out_dir.mkdir()

        # Try marker first (run as a fully isolated subprocess)
        try:
            result = subprocess.run(
                ["marker_single", str(pdf_path), str(out_dir)],
                capture_output=True,
                timeout=300,
            )
            if result.returncode != 0:
                logger.warning(
                    "marker exited with code %d: %s",
                    result.returncode,
                    result.stderr[:500] if result.stderr else "(no stderr)",
                )
                raise subprocess.CalledProcessError(result.returncode, "marker_single")
            # marker outputs <stem>/<stem>.md and images/
            md_candidates = list(out_dir.rglob("*.md"))
            markdown = md_candidates[0].read_text() if md_candidates else ""
            if not markdown.strip():
                logger.warning("marker produced empty output; falling back to OCR")
                raise ValueError("empty marker output")
        except Exception:
            logger.warning("marker not available or failed; falling back to OCR")
            markdown = _pdf_ocr_fallback(pdf_path)

        # Render page images using Pillow (requires pdf2image + poppler)
        try:
            page_images = _render_pdf_pages(pdf_path)
        except Exception:
            logger.warning("Page rendering failed", exc_info=True)
            page_images = []

        # Build page map (approximate: split markdown into equal-ish chunks)
        page_map: list[dict[str, Any]] = []
        if page_images:
            chunk_size = max(1, len(markdown) // len(page_images))
            for i in range(len(page_images)):
                start = i * chunk_size
                end = min((i + 1) * chunk_size, len(markdown))
                page_map.append({"page": i + 1, "char_start": start, "char_end": end})

    return page_images, markdown, page_map


def _render_pdf_pages(pdf_path: Path) -> list[bytes]:
    """Render PDF pages to PNG images. Requires pdf2image (poppler)."""
    try:
        from pdf2image import convert_from_path

        images = convert_from_path(str(pdf_path), dpi=150)
        result: list[bytes] = []
        for img in images:
            buf = io.BytesIO()
            img.save(buf, format="PNG")
            result.append(buf.getvalue())
        return result
    except ImportError:
        logger.warning("pdf2image not available — skipping page rendering")
        return []
    except Exception:
        logger.warning("Failed to render PDF pages", exc_info=True)
        return []


def _pdf_ocr_fallback(pdf_path: Path) -> str:
    """OCR each page of a PDF as a fallback."""
    try:
        from pdf2image import convert_from_path

        images = convert_from_path(str(pdf_path), dpi=200)
        parts: list[str] = []
        for i, img in enumerate(images, 1):
            text = pytesseract.image_to_string(img)
            parts.append(f"## Page {i}\n\n{text.strip()}\n")
        return "\n".join(parts)
    except Exception:
        logger.warning("OCR fallback failed for %s", pdf_path, exc_info=True)
        return "(Failed to extract text from PDF)"


# ── DOCX → markdown ──────────────────────────────────────────────────────────


async def docx_to_markdown(content: bytes) -> str:
    """Convert a DOCX file to markdown using pandoc."""
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, _docx_to_md_sync, content)


def _docx_to_md_sync(content: bytes) -> str:
    with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as tmp:
        tmp.write(content)
        tmp.flush()
        try:
            result = subprocess.run(
                ["pandoc", tmp.name, "-t", "markdown", "--wrap=none"],
                capture_output=True,
                text=True,
                timeout=120,
            )
            if result.returncode == 0:
                return result.stdout
            logger.warning("pandoc failed: %s", result.stderr)
            return f"(pandoc conversion failed: {result.stderr[:200]})"
        except FileNotFoundError:
            return "(pandoc not installed — cannot convert DOCX)"
        finally:
            Path(tmp.name).unlink(missing_ok=True)


# ── PPTX → markdown ──────────────────────────────────────────────────────────


async def pptx_to_markdown(content: bytes) -> str:
    """Convert a PPTX file to markdown using pandoc."""
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, _pptx_to_md_sync, content)


def _pptx_to_md_sync(content: bytes) -> str:
    with tempfile.NamedTemporaryFile(suffix=".pptx", delete=False) as tmp:
        tmp.write(content)
        tmp.flush()
        try:
            result = subprocess.run(
                ["pandoc", tmp.name, "-t", "markdown", "--wrap=none"],
                capture_output=True,
                text=True,
                timeout=120,
            )
            if result.returncode == 0:
                return result.stdout
            return f"(pandoc conversion failed: {result.stderr[:200]})"
        except FileNotFoundError:
            return "(pandoc not installed — cannot convert PPTX)"
        finally:
            Path(tmp.name).unlink(missing_ok=True)


# ── XLSX → markdown ──────────────────────────────────────────────────────────


async def xlsx_to_markdown(content: bytes) -> str:
    """Convert an XLSX file to markdown tables."""
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, _xlsx_to_md_sync, content)


def _xlsx_to_md_sync(content: bytes) -> str:
    try:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        parts: list[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"## {sheet_name}\n")
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                parts.append("(empty sheet)\n")
                continue
            # Header
            header = [str(c) if c is not None else "" for c in rows[0]]
            parts.append("| " + " | ".join(header) + " |")
            parts.append("| " + " | ".join(["---"] * len(header)) + " |")
            for row in rows[1:]:
                cells = [str(c) if c is not None else "" for c in row]
                parts.append("| " + " | ".join(cells) + " |")
            parts.append("")
        return "\n".join(parts)
    except ImportError:
        return "(openpyxl not installed — cannot convert XLSX)"
    except Exception as exc:
        return f"(XLSX conversion failed: {exc})"


# ── Image → markdown (OCR + description) ─────────────────────────────────────


async def image_to_markdown(content: bytes) -> str:
    """OCR an image and return the extracted text as markdown."""
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, _image_to_md_sync, content)


def _image_to_md_sync(content: bytes) -> str:
    try:
        img = Image.open(io.BytesIO(content))
        text = pytesseract.image_to_string(img)
        return f"## OCR Text\n\n{text.strip()}" if text.strip() else "(No text detected in image)"
    except Exception as exc:
        logger.warning("Image OCR failed: %s", exc)
        return f"(Image OCR failed: {exc})"


# ── Generic fallback ─────────────────────────────────────────────────────────


async def generic_to_text(content: bytes) -> str:
    """Best-effort text extraction — try UTF-8 decode."""
    try:
        return content.decode("utf-8")
    except UnicodeDecodeError:
        try:
            return content.decode("latin-1")
        except Exception:
            return "(Binary file — could not extract text)"
